import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

order_size = int(input("Podaj wielkość zamówienia- ile namiotów potrzebujesz?\n"))
order_time = int(input("Podaj tydzień odbioru zamowienia (od dzis, liczba)\n"))
supplies_stock = int(input("Podaj wstępny zapas magazynu: \n"))
assembling_time = int(input("Podaj czas potrzebny na złożenie przedmiotu: \n"))

weeks = []
for i in range(order_time):
    weeks.append(i+1)

level = 0
component = "tent"

record = []

for i in range(1,order_time+1):
    if i == 1:
        record.append(level)
    elif i == 2:
        record.append(component)
    else:
        record.append(0)

brutto = []
for i in range(1,order_time+1):
    if i < order_time:
        brutto.append(0)
    elif i == order_time:
        brutto.append(order_size)

stock = []
for i in range(1,order_time+1):
    stock.append(supplies_stock)

netto = []
for i in range(1,order_time+1):
    if i<order_time:
        netto.append(0)
    elif i == order_time:
        if order_size- supplies_stock>= 0:
            netto.append(order_size-supplies_stock)
        else:
            netto.append(0)          #potrzeba netto nie może być liczbą ujemną!

assembling = []
for i in range(1,order_time+1):
    if i != order_time-assembling_time:
        assembling.append(0)
    elif i == order_time-assembling_time:
        if order_size-supplies_stock>=0:
            assembling.append(order_size-supplies_stock)
        else:
            assembling.append(0)

pickup = []
for i in range(1,order_time+1):
    if i<order_time:
        pickup.append(0)
    elif i == order_time:
        pickup.append(order_size)

# dane
dane = {
    'Level': record,
    'Week': weeks,
    'BRUTTO': brutto,
    'Supplies stock': stock,
    'NETTO': netto,
    'Assembling': assembling,
    'Pickup': pickup
}

# Tworzenie ramki danych
df = pd.DataFrame(dane)

# Tworzenie pliku Excel
nazwa_pliku = 'tent.xlsx'
df.to_excel(nazwa_pliku, index=False)

print("Plik Excel został utworzony pomyślnie.")

wb = openpyxl.load_workbook(f"{nazwa_pliku}")
ws = wb["Sheet1"]

green = PatternFill(fill_type="solid", fgColor="8fbc8f")
beige = PatternFill(fill_type="solid", fgColor="FAF0E6")

for i in range(1,order_time+2):
    ws[f"B{i}"].fill = green
    #ws[f"B{i}"].aligment = Alignment(horizontal="center")
    wb.save(f"{nazwa_pliku}")

ws["A3"].fill = beige
wb.save(f"{nazwa_pliku}")

cells = ["A", "B", "C", "D", "E", "F", "G"]
'''
for cell in cells:
    for i in range(1, order_time+2):
        ws[f"{cell}{i}"].aligment = Alignment(horizontal="center")
    wb.save(f"{nazwa_pliku}")
'''