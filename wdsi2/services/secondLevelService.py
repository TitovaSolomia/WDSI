import webbrowser
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from file_management.convertation import dataframeToPdf, xlsxToDataFrame

def generateSecondLevelPdf(order_size, order_time, tent_time, supplies_stock, assembling_time, component, ordering_time):
    secondLevelAnalisePerform(order_size, order_time, tent_time, supplies_stock, assembling_time, component, ordering_time)

    file_path = component+".xlsx"  
    data_frame = xlsxToDataFrame(file_path)

    if data_frame is not None:
        print("DataFrame created successfully!")
        print(data_frame.head()) 
    else:
        print("Failed to create DataFrame.")

    file_path = "analise2.pdf"
    dataframeToPdf(data_frame, file_path)

    if file_path:
        webbrowser.open(file_path)


def secondLevelAnalisePerform(order_size,order_time, tent_time, supplies_stock, assembling_time, component, ordering_time):

    weeks = []
    for i in range(order_time):
        weeks.append(i+1)

    level = 2

    record = []

    for i in range(1,order_time+1):
        if i == 1:
            record.append(level)
        elif i == 2:
            record.append(component)
        else:
            record.append("-")

    amount = {
        "ropes": 6,
        "waterproof_material": 1, 
        "floor_material": 1, 
        "tubes": 5, 
        "herrings": 8, 
        "net": 1, 
        "lock": 1
    }

    brutto = []
    for i in range(1,order_time+1):
        if i != order_time-tent_time-assembling_time:
            brutto.append(0)
        elif i == order_time-tent_time-assembling_time:
            brutto.append(order_size*amount[f"{component}"])   
                                            
    stock = []
    for i in range(1,order_time+1):
        if i <= order_time-tent_time-assembling_time:
            stock.append(supplies_stock)
        if i > order_time-tent_time-assembling_time:
            if order_size*amount[f"{component}"]>=supplies_stock:
                stock.append(0)                                            
            elif order_size*amount[f"{component}"]<supplies_stock:
                stock.append(supplies_stock-order_size*amount[f"{component}"])    

    netto = []
    for i in range(1,order_time+1):
        if i != order_time-tent_time-assembling_time:
            netto.append(0)
        elif i == order_time-tent_time-assembling_time:
            if order_size*amount[f"{component}"]- supplies_stock>= 0:
                netto.append(order_size*amount[f"{component}"]-supplies_stock)
            else:
                netto.append(0)               #potrzeba netto nie może być liczbą ujemną!

    order = []
    for i in range(1,order_time+1):
        if i != order_time-tent_time-assembling_time-ordering_time:
            order.append(0)
        elif i == order_time-tent_time-assembling_time-ordering_time:
            if order_size*amount[f"{component}"]-supplies_stock>=0:
                order.append(order_size*amount[f"{component}"]-supplies_stock)
            else:
                order.append(0)

    pickup = []
    for i in range(1,order_time+1):
        if i != order_time-tent_time-assembling_time:
            pickup.append(0)
        elif i == order_time-tent_time-assembling_time:
            pickup.append(order_size*amount[f"{component}"])


    # dane
    dane = {
        'Level': record,
        'Week': weeks,
        'BRUTTO': brutto,
        'Supplies stock': stock,
        'NETTO': netto,
        'Order': order,
        'Pickup': pickup
    }

    # Tworzenie ramki danych
    df = pd.DataFrame(dane)

    # Tworzenie pliku Excel
    nazwa_pliku = f"{component}.xlsx"
    df.to_excel(nazwa_pliku, index=False)

    wb = openpyxl.load_workbook(f"{nazwa_pliku}")
    ws = wb["Sheet1"]

    green = PatternFill(fill_type="solid", fgColor="8fbc8f")
    beige = PatternFill(fill_type="solid", fgColor="FAF0E6")

    for i in range(1,order_time+2):
        ws[f"B{i}"].fill = green
        wb.save(f"{nazwa_pliku}")

    ws["A3"].fill = beige
    wb.save(f"{nazwa_pliku}")