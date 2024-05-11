import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

flag =True                                                                       #flaga informujaca o mozliwosci realizacji zamowienia

def record0(order_size,order_time,supplies_stock,tent_time):                     #funkcja do poziomu 0

    if order_time-tent_time<0:
        print("SORRY. THIS ORDER CANNOT BE COMPLETED IN THE TIME THAT YOU ENTERED.")
        global flag
        flag=False
    else:
        weeks = []
        for i in range(order_time):
            weeks.append(i+1)

        level = 0
        component = "tent"

        record = []
        for i in range(1,order_time+1):
            if i == 1:
                record.append(level)
            elif i == 2:
                record.append(component)
            else:
                record.append("")

        brutto = []
        for i in range(1,order_time+1):
            if i < order_time:
                brutto.append("")
            elif i == order_time:
                brutto.append(order_size)

        stock = []
        for i in range(1,order_time+1):
            stock.append(supplies_stock)

        netto = []
        for i in range(1,order_time+1):
            if i<order_time:
                netto.append("")
            elif i == order_time:
                if order_size- supplies_stock>= 0:
                    netto.append(order_size-supplies_stock)
                else:
                    netto.append(0)          #potrzeba netto nie może być liczbą ujemną!

        assembling = []
        for i in range(1,order_time+1):
            if i != order_time-tent_time:
                assembling.append("")
            elif i == order_time-tent_time:
                if order_size-supplies_stock>=0:
                    assembling.append(order_size-supplies_stock)
                else:
                    assembling.append("")

        pickup = []
        for i in range(1,order_time+1):
            if i<order_time:
                pickup.append("")
            elif i == order_time:
                pickup.append(order_size)

        # dane
        dane = {
            'Level': record,
            'Week': weeks,
            'BRUTTO': brutto,
            'Supplies stock': stock,
            'NETTO': netto,
            'Assembling': assembling,
            'Pickup': pickup
        }

        # Tworzenie ramki danych
        df = pd.DataFrame(dane)

        # Tworzenie pliku Excel
        nazwa_pliku = 'tent.xlsx'
        df.to_excel(nazwa_pliku, index=False)

        print("Excel file was created successfully.")

        wb = openpyxl.load_workbook(f"{nazwa_pliku}")
        ws = wb["Sheet1"]

        green = PatternFill(fill_type="solid", fgColor="8fbc8f")
        beige = PatternFill(fill_type="solid", fgColor="FAF0E6")

        for i in range(1,order_time+2):
            ws[f"B{i}"].fill = green
            wb.save(f"{nazwa_pliku}")

        ws["A3"].fill = beige
        wb.save(f"{nazwa_pliku}")


def record1(order_size,order_time,supplies_stock,assembling_time, component):             #funkcja do poziomu pierwszego!

    if order_time-tent_time-assembling_time<0:
        print("SORRY. THIS ORDER CANNOT BE COMPLETED IN THE TIME THAT YOU ENTERED.")
        global flag
        flag=False
    else:
        weeks = []
        for i in range(order_time):
            weeks.append(i+1)

        level = 1

        record = []

        for i in range(1,order_time+1):
            if i == 1:
                record.append(level)
            elif i == 2:
                record.append(component)
            else:
                record.append("")

        brutto = []
        for i in range(1,order_time+1):
            if i != order_time-tent_time:
                brutto.append("")
            elif i == order_time-tent_time:
                brutto.append(order_size)          #potrzebujemy po 1 z kazdych z tych elementow do jednego gotowego namiotu, stad liczba potrzebnych
                                                #produktow brutto rowna liczbie potrzebnego calego zamowienia
        stock = []
        for i in range(1,order_time+1):
            if i <= order_time-tent_time:
                stock.append(supplies_stock)
            if i > order_time-tent_time:
                if order_size>=supplies_stock:
                    stock.append(0)
                elif order_size<supplies_stock:
                    stock.append(supplies_stock-order_size)     

        netto = []
        for i in range(1,order_time+1):
            if i != order_time-tent_time:
                netto.append("")
            elif i == order_time-tent_time:
                if order_size- supplies_stock>= 0:
                    netto.append(order_size-supplies_stock)
                else:
                    netto.append(0)               #potrzeba netto nie może być liczbą ujemną!

        assembling = []
        for i in range(1,order_time+1):
            if i != order_time-tent_time-assembling_time:
                assembling.append("")
            elif i == order_time-tent_time-assembling_time:
                if order_size-supplies_stock>=0:
                    assembling.append(order_size-supplies_stock)
                else:
                    assembling.append("")

        pickup = []
        for i in range(1,order_time+1):
            if i != order_time-tent_time:
                pickup.append("")
            elif i == order_time-tent_time:
                pickup.append(order_size)

        # dane
        dane = {
            'Level': record,
            'Week': weeks,
            'BRUTTO': brutto,
            'Supplies stock': stock,
            'NETTO': netto,
            'Assembling': assembling,
            'Pickup': pickup
        }

        # Tworzenie ramki danych
        df = pd.DataFrame(dane)

        # Tworzenie pliku Excel
        nazwa_pliku = f"{component}.xlsx"
        df.to_excel(nazwa_pliku, index=False)

        wb = openpyxl.load_workbook(f"{nazwa_pliku}")
        ws = wb["Sheet1"]

        #Tworzenie kolorów 
        green = PatternFill(fill_type="solid", fgColor="8fbc8f")
        beige = PatternFill(fill_type="solid", fgColor="FAF0E6")

        #Użycie kolorów w komórkach tabeli
        for i in range(1,order_time+2):
            ws[f"B{i}"].fill = green
            wb.save(f"{nazwa_pliku}")

        ws["A3"].fill = beige
        wb.save(f"{nazwa_pliku}")

#---------------------------------------------------------------------------------------------------

def record2(order_size,order_time,supplies_stock2,ordering_time, component2, assembling_time2):            #fukcja do poziomu drugiego!

    if order_time-tent_time-assembling_time2-ordering_time<0:
        print("SORRY. THIS ORDER CANNOT BE COMPLETED IN THE TIME THAT YOU ENTERED.")
    else:
        weeks = []
        for i in range(order_time):
            weeks.append(i+1)

        level = 2

        record = []
        for i in range(1,order_time+1):
            if i == 1:
                record.append(level)
            elif i == 2:
                record.append(component2)
            else:
                record.append("")
        
        #Potrzebna ilosc poszczegolnych czesci skladowych do zlozenia 1 namiotu
        amount = {
            "ropes": 6,
            "waterproof_material": 1, 
            "floor_material": 1, 
            "tubes": 5, 
            "herrings": 8, 
            "net": 1, 
            "lock": 1
        }

        brutto = []
        for i in range(1,order_time+1):
            if i != order_time-tent_time-assembling_time2:
                brutto.append("")
            elif i == order_time-tent_time-assembling_time2:
                brutto.append(order_size*amount[f"{component2}"])  
                                                
        stock = []
        for i in range(1,order_time+1):
            if i <= order_time-tent_time-assembling_time2:
                stock.append(supplies_stock2)
            if i > order_time-tent_time-assembling_time2:
                if order_size*amount[f"{component2}"]>=supplies_stock2:
                    stock.append(0)                                            
                elif order_size*amount[f"{component2}"]<supplies_stock2:
                    stock.append(supplies_stock2-order_size*amount[f"{component2}"])   

        netto = []
        for i in range(1,order_time+1):
            if i != order_time-tent_time-assembling_time2:
                netto.append("")
            elif i == order_time-tent_time-assembling_time2:
                if order_size*amount[f"{component2}"]- supplies_stock2>= 0:
                    netto.append(order_size*amount[f"{component2}"]-supplies_stock2)
                else:
                    netto.append(0)               #potrzeba netto nie może być liczbą ujemną!

        order = []
        for i in range(1,order_time+1):
            if i != order_time-tent_time-assembling_time2-ordering_time:
                order.append("")
            elif i == order_time-tent_time-assembling_time2-ordering_time:
                if order_size*amount[f"{component2}"]-supplies_stock2>0:
                    order.append(order_size*amount[f"{component2}"]-supplies_stock2)
                else:
                    order.append("")

        pickup = []
        for i in range(1,order_time+1):
            if i != order_time-tent_time-assembling_time2:
                pickup.append("")
            elif i == order_time-tent_time-assembling_time2:
                pickup.append(order_size*amount[f"{component2}"])

        # dane
        dane = {
            'Level': record,
            'Week': weeks,
            'BRUTTO': brutto,
            'Supplies stock': stock,
            'NETTO': netto,
            'Order': order,
            'Pickup': pickup
        }

        # Tworzenie ramki danych
        df = pd.DataFrame(dane)

        # Tworzenie pliku Excel
        nazwa_pliku = f"{component2}.xlsx"
        df.to_excel(nazwa_pliku, index=False)

        wb = openpyxl.load_workbook(f"{nazwa_pliku}")
        ws = wb["Sheet1"]

        #Tworzenie kolorów 
        green = PatternFill(fill_type="solid", fgColor="8fbc8f")
        beige = PatternFill(fill_type="solid", fgColor="FAF0E6")

        #Użycie kolorów w komórkach tabeli
        for i in range(1,order_time+2):
            ws[f"B{i}"].fill = green
            wb.save(f"{nazwa_pliku}")

        ws["A3"].fill = beige
        wb.save(f"{nazwa_pliku}")



print("Material Requirements Planning System for a shop with tents\n")
print("Level 0- tent")

order_size = int(input("Enter the order size- how many tents do you need?\n"))
order_time = int(input("Enter the week in which order (tents) should be ready to pickup (from today, the number):\n"))
supplies_stock0 = int(input("Enter the initial stock of ready-made tents: \n"))
tent_time = int(input("Enter the time (in weeks) needed to assemble the tent from the ready-made parts: \n"))

record0(order_size,order_time,supplies_stock0,tent_time)

if flag:
    components = ["roof","floor", "frame", "entry"]                                                 #POZIOM 1
    time2_arr = []

    print("Level 1: \n")

    for component in components:

        print(f"{component}")
        supplies_stock = int(input("Enter the initial stock of the item: \n"))                             #inny dla kazdej skladowej 
        assembling_time = int(input("Enter the time (in weeks) needed to assemble the item: \n"))
        time2_arr.append(assembling_time)

        record1(order_size,order_time,supplies_stock,assembling_time, component)

        print("Excel files were created successfully.")

if flag:
    components2 = ["ropes","waterproof_material", "floor_material", "tubes", "herrings", "net", "lock"]
    print("Level 2: \n")

    for component2 in components2:                                                                   #POZIOM 2
        
        print(f"{component2}")
        supplies_stock2 = int(input("Enter the initial stock of the item: \n"))                             #inny dla kazdej skladowej 
        ordering_time = int(input("Enter the time (in weeks) needed to order the item: \n"))

        if component2 == "ropes":
            if time2_arr[0] >= time2_arr[3]:
                assembling_time2 = time2_arr[0]
            else:
                assembling_time2 = time2_arr[3]
        elif component2 == "waterproof_material":
            assembling_time2 = time2_arr[0]
        elif component2 == "floor_material":
            assembling_time2 = time2_arr[1]
        elif component2 == "tubes":
            assembling_time2 = time2_arr[2]
        elif component2 == "herrings":
            assembling_time2 = time2_arr[2]
        elif component2 == "net":
            assembling_time2 = time2_arr[3]
        elif component2 == "lock":
            assembling_time2 = time2_arr[3]

        record2(order_size,order_time,supplies_stock2,ordering_time, component2,assembling_time2)

        print("Excel files were created successfully.")